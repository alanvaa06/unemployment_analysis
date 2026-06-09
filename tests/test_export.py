"""Test the multi-sheet Excel export (one sheet per chart's data)."""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

CACHE = Path("data/cache")


@pytest.mark.skipif(not (CACHE / "observations.parquet").exists(),
                    reason="needs a populated cache")
def test_build_workbook_has_sheet_per_chart(tmp_path) -> None:
    from dashboard.export import build_workbook
    out = tmp_path / "analysis.xlsx"
    build_workbook(out, base="2022-11", compare="2026-05", cache_dir=CACHE)
    assert out.exists()
    wb = openpyxl.load_workbook(out, read_only=True)
    names = set(wb.sheetnames)
    # a sheet for each major chart
    for expected in ["Waterfall", "IndustryChange", "FreezeVsCuts", "StateDispersion",
                     "ChangeDistribution", "StatesSwarm", "Beveridge", "Occupations"]:
        assert expected in names, f"missing sheet {expected}"
    # sheets carry data (header + rows)
    ws = wb["Waterfall"]
    assert ws.max_row >= 2 and ws.max_column >= 2
    wb.close()
